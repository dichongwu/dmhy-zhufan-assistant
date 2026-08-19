#!/usr/bin/env python3
"""Scrape, validate, export, and organize a DMHY team's complete topic listing."""

from __future__ import annotations

import argparse
import csv
import html
import json
import os
import re
import subprocess
import sys
import tempfile
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from datetime import datetime, timedelta, timezone
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urljoin, urlparse


SITE_URL = "https://share.dmhy.org"
LIST_PATH_RE = re.compile(r"^/topics/list/(team_id|user_id)/(\d+)(?:/page/\d+)?/?$")
DATE_RE = re.compile(r"(\d{4})/(\d{2})/(\d{2})\s+(\d{2}):(\d{2})")
TOPIC_ID_RE = re.compile(r"/topics/view/(\d+)(?:_|\.html)")
RAW_FIELDS = [
    "topic_id",
    "published_at",
    "category",
    "title",
    "detail_url",
    "magnet",
    "size",
    "seeders",
    "downloads",
    "completed",
    "publisher",
    "source_page",
]
TECHNICAL_TOKEN_RE = re.compile(
    r"(?:web-?rip|bd-?rip|blu-?ray|hevc|x26[45]|avc|aac|flac|1080|720p|2160|"
    r"remux|ma10p|mp4|mkv|10bit|hi10|字幕|内封|內封|外挂|外掛|内挂|內掛|"
    r"无字|無字|中字|简体|繁体|簡體|繁體|英文|英语|英語|ass|srt|^fin$|^end$)",
    re.IGNORECASE,
)
BATCH_TOKEN_RE = re.compile(
    r"^\s*\d{1,3}(?:\s*[-~+]\s*\d{1,3})?(?:\s*(?:修正|精校|重制|重製|合集|全))*\s*$",
    re.IGNORECASE,
)
EPISODE_TOKEN_RE = re.compile(
    r"^(?:EP\s*)?\d{1,3}(?:\.\d+)?(?:v\d+)?(?:\(\d+\))?$|"
    r"^(?:OVA|OAD|SP|NCOP|NCED|OP|ED|PV|CM)\d*$|^MAD$",
    re.IGNORECASE,
)
TRAILING_EPISODE_RE = re.compile(
    r"\s+-\s+((?:EP\s*)?\d{1,3}(?:\.\d+)?(?:v\d+)?(?:\(\d+\))?|"
    r"(?:OVA|OAD|SP|NCOP|NCED|OP|ED|PV|CM)\d*|MAD)\s*$|"
    r"\s+(\d{1,3}\s*[-~+]\s*\d{1,3})\s*$",
    re.IGNORECASE,
)
GROUPED_FIELDS = [
    "series_rank",
    "series_name",
    "series_latest_at",
    "release_count",
    "episode_or_batch",
    "published_at",
    "size",
    "category",
    "publisher",
    "title",
    "detail_url",
    "magnet",
    "topic_id",
    "source_page",
]
NEW_ITEM_FILL_COLOR = "C6EFCE"


def clean_text(parts: list[str]) -> str:
    return " ".join(html.unescape("".join(parts)).split())


class TopicTableParser(HTMLParser):
    def __init__(self, page: int) -> None:
        super().__init__(convert_charrefs=True)
        self.page = page
        self.in_title = False
        self.title_parts: list[str] = []
        self.in_table = False
        self.in_body = False
        self.current_row: list[dict] | None = None
        self.current_cell: dict | None = None
        self.current_link: dict | None = None
        self.rows: list[dict] = []

    @property
    def page_title(self) -> str:
        return clean_text(self.title_parts)

    def handle_starttag(self, tag: str, attrs_list: list[tuple[str, str | None]]) -> None:
        attrs = dict(attrs_list)
        if tag == "title":
            self.in_title = True
        elif tag == "table" and attrs.get("id") == "topic_list":
            self.in_table = True
        elif self.in_table and tag == "tbody":
            self.in_body = True
        elif self.in_body and tag == "tr":
            self.current_row = []
        elif self.current_row is not None and tag == "td":
            self.current_cell = {"text": [], "links": []}
        elif self.current_cell is not None and tag == "a":
            self.current_link = {"attrs": attrs, "text": []}

    def handle_data(self, data: str) -> None:
        if self.in_title:
            self.title_parts.append(data)
        if self.current_cell is not None:
            self.current_cell["text"].append(data)
        if self.current_link is not None:
            self.current_link["text"].append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag == "title":
            self.in_title = False
        elif tag == "a" and self.current_link is not None:
            self.current_link["text"] = clean_text(self.current_link["text"])
            self.current_cell["links"].append(self.current_link)
            self.current_link = None
        elif tag == "td" and self.current_cell is not None:
            self.current_cell["text"] = clean_text(self.current_cell["text"])
            self.current_row.append(self.current_cell)
            self.current_cell = None
        elif tag == "tr" and self.current_row is not None:
            topic = self._make_topic(self.current_row)
            if topic:
                self.rows.append(topic)
            self.current_row = None
        elif tag == "tbody" and self.in_body:
            self.in_body = False
        elif tag == "table" and self.in_table:
            self.in_table = False

    def _make_topic(self, cells: list[dict]) -> dict | None:
        if len(cells) < 9:
            return None

        title_link = next(
            (link for link in cells[2]["links"] if link["attrs"].get("href", "").startswith("/topics/view/")),
            None,
        )
        if not title_link:
            return None

        detail_path = title_link["attrs"]["href"]
        topic_id_match = TOPIC_ID_RE.search(detail_path)
        date_match = DATE_RE.search(cells[0]["text"])
        if not topic_id_match or not date_match:
            return None

        magnet = next(
            (
                link["attrs"]["data-magnet"]
                for link in cells[3]["links"]
                if link["attrs"].get("data-magnet", "").startswith("magnet:")
            ),
            "",
        )
        if not magnet:
            full_magnet = next(
                (
                    link["attrs"]["href"]
                    for link in cells[3]["links"]
                    if link["attrs"].get("href", "").startswith("magnet:")
                ),
                "",
            )
            magnet = full_magnet.split("&", 1)[0]

        published_at = "-".join(date_match.group(1, 2, 3)) + " " + ":".join(date_match.group(4, 5))
        return {
            "topic_id": int(topic_id_match.group(1)),
            "published_at": published_at,
            "category": cells[1]["text"],
            "title": title_link["text"],
            "detail_url": urljoin(SITE_URL, detail_path),
            "magnet": magnet,
            "size": cells[4]["text"],
            "seeders": cells[5]["text"],
            "downloads": cells[6]["text"],
            "completed": cells[7]["text"],
            "publisher": cells[8]["text"],
            "source_page": self.page,
        }


def normalize_source(value: str) -> tuple[int, str]:
    if value.isdigit():
        source_id = int(value)
        source_kind = "team_id"
    else:
        parsed = urlparse(value)
        if parsed.scheme not in {"http", "https"} or parsed.hostname != "share.dmhy.org":
            raise ValueError("source must be a share.dmhy.org team/user URL or numeric team ID")
        match = LIST_PATH_RE.match(parsed.path)
        if not match:
            raise ValueError(
                "URL must look like https://share.dmhy.org/topics/list/team_id/657 "
                "or https://share.dmhy.org/topics/list/user_id/759200"
            )
        source_kind = match.group(1)
        source_id = int(match.group(2))
    if source_id < 1:
        raise ValueError("source ID must be positive")
    return source_id, f"{SITE_URL}/topics/list/{source_kind}/{source_id}"


def page_url(list_url: str, page: int) -> str:
    return list_url if page == 1 else f"{list_url}/page/{page}"


def fetch_page(list_url: str, page: int) -> str:
    command = [
        "curl",
        "-L",
        "--compressed",
        "--fail",
        "--silent",
        "--show-error",
        "--retry",
        "4",
        "--retry-all-errors",
        "--connect-timeout",
        "15",
        "--max-time",
        "90",
        "-A",
        "Mozilla/5.0 (compatible; DMHYTeamExporter/1.0)",
        page_url(list_url, page),
    ]
    resolve_ip = os.environ.get("DMHY_RESOLVE_IP", "").strip()
    if resolve_ip:
        if not re.fullmatch(r"(?:\d{1,3}\.){3}\d{1,3}", resolve_ip):
            raise ValueError("DMHY_RESOLVE_IP must be an IPv4 address")
        command[1:1] = ["--resolve", f"share.dmhy.org:443:{resolve_ip}"]
    last_error = ""
    for attempt in range(1, 4):
        result = subprocess.run(command, capture_output=True)
        if result.returncode == 0:
            return result.stdout.decode("utf-8", errors="replace")
        last_error = result.stderr.decode("utf-8", errors="replace").strip()
        time.sleep(attempt * 2)
    raise RuntimeError(f"page {page} failed after retries: {last_error}")


def parse_page(page: int, source: str) -> TopicTableParser:
    parser = TopicTableParser(page)
    parser.feed(source)
    if "動漫花園" not in parser.page_title:
        raise RuntimeError(f"page {page} did not return a recognizable DMHY page")
    return parser


def load_page(list_url: str, page: int) -> TopicTableParser:
    return parse_page(page, fetch_page(list_url, page))


def discover_last_page(list_url: str, cache: dict[int, TopicTableParser]) -> int:
    low = 1
    high = 2
    while high <= 4096:
        parsed = load_page(list_url, high)
        cache[high] = parsed
        if not parsed.rows:
            break
        low = high
        high *= 2
    else:
        raise RuntimeError("could not find the last page below page 4096")

    while low + 1 < high:
        middle = (low + high) // 2
        parsed = load_page(list_url, middle)
        cache[middle] = parsed
        if parsed.rows:
            low = middle
        else:
            high = middle
    return low


def scrape(
    list_url: str, last_page: int | None, workers: int, since_days: int | None = None
) -> tuple[list[dict], int, str]:
    first_page = load_page(list_url, 1)
    if not first_page.rows:
        raise RuntimeError("the listing contains no topic rows")
    cache = {1: first_page}

    cutoff = None
    if since_days is not None:
        requested_last_page = last_page
        cutoff = (datetime.now().astimezone() - timedelta(days=since_days)).strftime("%Y-%m-%d %H:%M")
        page = 1
        while cache[page].rows[-1]["published_at"] >= cutoff:
            page += 1
            if requested_last_page is not None and page > requested_last_page:
                page = requested_last_page
                break
            if page > 4096:
                raise RuntimeError("could not find a page older than the requested date window")
            parsed = load_page(list_url, page)
            if not parsed.rows:
                page -= 1
                break
            cache[page] = parsed
        last_page = page
        print(f"Using {last_page} pages for releases since {cutoff}", flush=True)
    elif last_page is None:
        print("Discovering the last page...", flush=True)
        last_page = discover_last_page(list_url, cache)
        print(f"Found {last_page} pages", flush=True)

    page_results = {page: parsed.rows for page, parsed in cache.items() if page <= last_page and parsed.rows}
    missing_pages = [page for page in range(1, last_page + 1) if page not in page_results]
    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(load_page, list_url, page): page for page in missing_pages}
        completed_pages = len(page_results)
        for future in as_completed(futures):
            page = futures[future]
            rows = future.result().rows
            if not rows:
                raise RuntimeError(f"page {page} contained no topic rows")
            page_results[page] = rows
            completed_pages += 1
            print(f"\rFetched {completed_pages}/{last_page} pages", end="", flush=True)
    if missing_pages:
        print()

    latest_page = load_page(list_url, 1)
    initial_ids = [row["topic_id"] for row in first_page.rows]
    latest_ids = [row["topic_id"] for row in latest_page.rows]
    if initial_ids != latest_ids:
        raise RuntimeError("the listing changed during the crawl; rerun to avoid a shifted-page snapshot")

    unique: dict[int, dict] = {}
    raw_count = 0
    for page in range(1, last_page + 1):
        for row in page_results[page]:
            raw_count += 1
            unique[row["topic_id"]] = row
    if len(unique) != raw_count:
        raise RuntimeError(f"found {raw_count - len(unique)} duplicate topic IDs across pages; rerun the crawl")

    rows = sorted(unique.values(), key=lambda row: (row["published_at"], row["topic_id"]), reverse=True)
    if cutoff is not None:
        rows = [row for row in rows if row["published_at"] >= cutoff]
        if not rows:
            raise RuntimeError("the requested date window contains no releases")
    team_name_match = re.match(r"(.+?)\s+\(#\d+\)", first_page.page_title)
    source_kind = "user" if "/user_id/" in list_url else "team"
    team_name = team_name_match.group(1) if team_name_match else f"{source_kind}-{list_url.rsplit('/', 1)[-1]}"
    return rows, last_page, team_name


def validate_rows(rows: list[dict]) -> None:
    required = ("topic_id", "published_at", "category", "title", "detail_url", "magnet", "size", "publisher")
    for index, row in enumerate(rows, start=1):
        missing = [field for field in required if not row.get(field)]
        if missing:
            raise RuntimeError(f"row {index} is missing required fields: {', '.join(missing)}")
        if not row["magnet"].startswith("magnet:?xt=urn:btih:"):
            raise RuntimeError(f"row {index} has an invalid magnet URI")


def safe_basename(value: str) -> str:
    name = re.sub(r"[^\w.-]+", "-", value.casefold(), flags=re.UNICODE).strip("-.")
    if not name or name in {".", ".."}:
        raise ValueError("could not derive a safe output name")
    return name


def write_raw_outputs(
    rows: list[dict], output_dir: Path, basename: str, list_url: str, pages: int, team_name: str
) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / f"{basename}.csv"
    json_path = output_dir / f"{basename}.json"

    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=RAW_FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    payload = {
        "source": list_url,
        "team": team_name,
        "scraped_at": datetime.now(timezone.utc).isoformat(),
        "pages": pages,
        "count": len(rows),
        "items": rows,
    }
    with json_path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")
    return csv_path, json_path


def is_episode_or_batch(token: str) -> bool:
    return bool(
        BATCH_TOKEN_RE.fullmatch(token)
        or EPISODE_TOKEN_RE.fullmatch(token)
        or re.search(r"(?:合集|全\s*\d+\s*集|Anime MV|OVA|OAD|Special|特典)", token, re.IGNORECASE)
    )


def parse_series_title(title: str) -> tuple[str, str]:
    normalized = title.replace("\u200b", "").replace("\ufeff", "").strip()
    group_match = re.match(r"^\[([^\]]+)\]\s*(.*)$", normalized)
    rest = group_match.group(2).strip() if group_match else normalized

    # Early uploads use [group][Chinese title][English title][episode][technical info].
    if rest.startswith("["):
        title_tokens: list[str] = []
        episode = ""
        for token in re.findall(r"\[([^\]]+)\]", rest):
            token = token.strip()
            if TECHNICAL_TOKEN_RE.search(token):
                break
            if title_tokens and is_episode_or_batch(token):
                episode = token
                break
            title_tokens.append(token)
        series = " / ".join(dict.fromkeys(title_tokens)).strip()
        return series or rest, episode or "完整作品"

    boundary = len(rest)
    episode = ""
    for bracket in re.finditer(r"\[([^\]]+)\]", rest):
        token = bracket.group(1).strip()
        if TECHNICAL_TOKEN_RE.search(token) or is_episode_or_batch(token):
            boundary = bracket.start()
            if is_episode_or_batch(token):
                episode = token
            break

    core = rest[:boundary].strip()
    # Some groups, notably 7³ACG, separate the series name from a batch or
    # movie-part marker with a vertical bar: "Series | 01-12+SPx2".
    if "|" in core:
        pipe_series, pipe_marker = core.rsplit("|", 1)
        if pipe_series.strip():
            core = pipe_series.strip()
            episode = pipe_marker.strip() or episode

    if not episode:
        episode_match = TRAILING_EPISODE_RE.search(core)
        if episode_match:
            episode = episode_match.group(1) or episode_match.group(2)
            core = core[: episode_match.start()].strip()
    return core or rest, episode or "完整作品"


def load_export(source: Path) -> tuple[str, list[dict]]:
    with source.open(encoding="utf-8") as handle:
        payload = json.load(handle)
    items = payload.get("items")
    if not isinstance(items, list) or not items:
        raise ValueError("source JSON does not contain a non-empty items array")
    fallback_name = source.stem.removesuffix("_topics")
    team_name = str(payload.get("team") or fallback_name or "DMHY")
    validate_rows(items)
    return team_name, items


def group_items(items: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for item in items:
        series_name, episode = parse_series_title(item["title"])
        enriched = dict(item)
        enriched["series_name"] = series_name
        enriched["episode_or_batch"] = episode
        grouped[series_name].append(enriched)

    groups = []
    for series_name, releases in grouped.items():
        releases.sort(key=lambda item: (item["published_at"], item["topic_id"]), reverse=True)
        groups.append({"series_name": series_name, "latest_at": releases[0]["published_at"], "releases": releases})
    groups.sort(key=lambda group: (group["latest_at"], group["series_name"]), reverse=True)
    return groups


def write_grouped_csv(groups: list[dict], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=GROUPED_FIELDS)
        writer.writeheader()
        for rank, group in enumerate(groups, start=1):
            releases = group["releases"]
            for release in releases:
                writer.writerow(
                    {
                        "series_rank": rank,
                        "series_name": group["series_name"],
                        "series_latest_at": group["latest_at"],
                        "release_count": len(releases),
                        "episode_or_batch": release["episode_or_batch"],
                        "published_at": release["published_at"],
                        "size": release["size"],
                        "category": release["category"],
                        "publisher": release["publisher"],
                        "title": release["title"],
                        "detail_url": release["detail_url"],
                        "magnet": release["magnet"],
                        "topic_id": release["topic_id"],
                        "source_page": release["source_page"],
                    }
                )


def escaped(value: object) -> str:
    return html.escape(str(value), quote=True)


def render_group(group: dict, rank: int) -> str:
    search_text = " ".join([group["series_name"]] + [release["title"] for release in group["releases"]]).casefold()
    rows = []
    for release in group["releases"]:
        rows.append(
            "<tr>"
            f"<td class='episode'>{escaped(release['episode_or_batch'])}</td>"
            f"<td class='date'>{escaped(release['published_at'])}</td>"
            f"<td class='size'>{escaped(release['size'])}</td>"
            f"<td>{escaped(release['category'])}</td>"
            f"<td>{escaped(release['publisher'])}</td>"
            f"<td class='release-title'>{escaped(release['title'])}</td>"
            "<td class='links'>"
            f"<a href='{escaped(release['detail_url'])}' target='_blank' rel='noreferrer'>详情</a>"
            f"<a href='{escaped(release['magnet'])}'>磁链</a>"
            "</td></tr>"
        )
    open_attribute = " open" if rank <= 12 else ""
    return (
        f"<details class='series' data-search='{escaped(search_text)}'{open_attribute}>"
        "<summary>"
        f"<span class='rank'>{rank}</span>"
        f"<span class='series-name'>{escaped(group['series_name'])}</span>"
        f"<span class='series-meta'>{escaped(group['latest_at'])} · {len(group['releases'])} 条</span>"
        "</summary><div class='table-wrap'><table>"
        "<thead><tr><th>集数/批次</th><th>发布日期</th><th>大小</th><th>分类</th><th>发布者</th><th>发布标题</th><th>链接</th></tr></thead>"
        f"<tbody>{''.join(rows)}</tbody></table></div></details>"
    )


def write_grouped_html(team_name: str, groups: list[dict], item_count: int, output_path: Path) -> None:
    newest = groups[0]["latest_at"]
    oldest = min(release["published_at"] for group in groups for release in group["releases"])
    sections = "".join(render_group(group, rank) for rank, group in enumerate(groups, start=1))
    report_title = f"{team_name} 按片名整理"
    document = f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{escaped(report_title)}</title>
<style>
:root {{ color-scheme:light; --ink:#17201c; --muted:#667069; --line:#d8ded9; --soft:#f3f6f4; --accent:#176b46; --link:#0b5cad; }}
* {{ box-sizing:border-box; }}
body {{ margin:0; color:var(--ink); background:#fff; font:14px/1.5 -apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif; letter-spacing:0; }}
header {{ position:sticky; top:0; z-index:10; border-bottom:1px solid var(--line); background:rgba(255,255,255,.96); backdrop-filter:blur(10px); }}
.header-inner, main {{ width:min(1600px,calc(100% - 32px)); margin:0 auto; }}
.header-inner {{ display:grid; grid-template-columns:minmax(220px,1fr) minmax(260px,520px); gap:24px; align-items:center; padding:16px 0; }}
h1 {{ margin:0; font-size:22px; line-height:1.25; letter-spacing:0; }}
.summary {{ margin-top:4px; color:var(--muted); font-size:13px; }}
input {{ width:100%; height:40px; border:1px solid #aeb8b1; border-radius:6px; padding:0 12px; background:#fff; color:var(--ink); font:inherit; letter-spacing:0; outline:none; }}
input:focus {{ border-color:var(--accent); box-shadow:0 0 0 3px rgba(23,107,70,.12); }}
main {{ padding:18px 0 48px; }} .result-count {{ margin:0 0 12px; color:var(--muted); }}
.series {{ border-top:1px solid var(--line); }} .series:last-child {{ border-bottom:1px solid var(--line); }} .series[hidden] {{ display:none; }}
summary {{ display:grid; grid-template-columns:42px minmax(0,1fr) auto; gap:12px; align-items:center; min-height:56px; padding:9px 10px 9px 4px; cursor:pointer; list-style:none; }}
summary::-webkit-details-marker {{ display:none; }} summary:hover {{ background:var(--soft); }}
.rank {{ color:var(--muted); text-align:right; font-variant-numeric:tabular-nums; }}
.series-name {{ min-width:0; font-size:16px; font-weight:650; overflow-wrap:anywhere; }}
.series-meta {{ color:var(--muted); white-space:nowrap; font-variant-numeric:tabular-nums; }}
.table-wrap {{ overflow-x:auto; border-top:1px solid var(--line); }}
table {{ width:100%; min-width:1060px; border-collapse:collapse; table-layout:fixed; }}
th,td {{ border-bottom:1px solid #e6eae7; padding:8px 10px; text-align:left; vertical-align:top; overflow-wrap:anywhere; }}
th {{ background:var(--soft); color:#465049; font-size:12px; font-weight:650; }}
th:nth-child(1) {{ width:90px; }} th:nth-child(2) {{ width:138px; }} th:nth-child(3) {{ width:86px; }} th:nth-child(4) {{ width:72px; }} th:nth-child(5) {{ width:110px; }} th:nth-child(7) {{ width:104px; }}
.episode {{ font-weight:700; color:var(--accent); }} .date,.size {{ white-space:nowrap; font-variant-numeric:tabular-nums; }}
.release-title {{ color:#38413b; }} .links {{ white-space:nowrap; }}
a {{ color:var(--link); text-decoration:none; }} a:hover {{ text-decoration:underline; }} .links a + a {{ margin-left:12px; }}
@media (max-width:720px) {{
  .header-inner,main {{ width:min(100% - 20px,1600px); }} .header-inner {{ grid-template-columns:1fr; gap:10px; padding:12px 0; }}
  h1 {{ font-size:19px; }} summary {{ grid-template-columns:32px minmax(0,1fr); gap:9px; }} .series-meta {{ grid-column:2; white-space:normal; }}
}}
</style>
</head>
<body>
<header><div class="header-inner">
  <div><h1>{escaped(report_title)}</h1><div class="summary">{len(groups):,} 个片名 · {item_count:,} 条发布 · {escaped(oldest)} 至 {escaped(newest)}</div></div>
  <input id="search" type="search" placeholder="搜索片名或发布标题" autocomplete="off">
</div></header>
<main><p class="result-count" id="resultCount">显示 {len(groups):,} 个片名</p><div id="seriesList">{sections}</div></main>
<script>
const input=document.getElementById('search');
const groups=[...document.querySelectorAll('.series')];
const count=document.getElementById('resultCount');
input.addEventListener('input',()=>{{
  const query=input.value.trim().toLocaleLowerCase(); let visible=0;
  for(const group of groups){{ const match=!query||group.dataset.search.includes(query); group.hidden=!match; if(match)visible+=1; }}
  count.textContent=`显示 ${{visible.toLocaleString()}} 个片名`;
}});
</script>
</body>
</html>
"""
    output_path.write_text(document, encoding="utf-8")


def write_grouped_outputs(
    team_name: str, items: list[dict], output_dir: Path, basename: str
) -> tuple[Path, Path, int]:
    groups = group_items(items)
    csv_path = output_dir / f"{basename}.csv"
    html_path = output_dir / f"{basename}.html"
    write_grouped_csv(groups, csv_path)
    write_grouped_html(team_name, groups, len(items), html_path)
    return csv_path, html_path, len(groups)


def read_xlsx_groups(sheet) -> list[dict]:
    groups: list[dict] = []
    current: dict | None = None
    for row in range(2, sheet.max_row + 1):
        dimension = sheet.row_dimensions[row]
        series_name = sheet.cell(row, 1).value
        if dimension.collapsed and series_name:
            current = {"series_name": str(series_name), "latest_at": "", "releases": []}
            groups.append(current)
            continue
        if current is None or not dimension.hidden:
            continue
        detail_cell = sheet.cell(row, 8)
        magnet_cell = sheet.cell(row, 9)
        detail_url = detail_cell.hyperlink.target if detail_cell.hyperlink else str(detail_cell.value or "")
        magnet = magnet_cell.hyperlink.target if magnet_cell.hyperlink else str(magnet_cell.value or "")
        topic_match = TOPIC_ID_RE.search(detail_url)
        release = {
            "topic_id": int(topic_match.group(1)) if topic_match else 0,
            "published_at": str(sheet.cell(row, 3).value or ""),
            "category": str(sheet.cell(row, 5).value or ""),
            "title": str(sheet.cell(row, 7).value or ""),
            "detail_url": detail_url,
            "magnet": magnet,
            "size": str(sheet.cell(row, 4).value or ""),
            "seeders": "",
            "downloads": "",
            "completed": "",
            "publisher": str(sheet.cell(row, 6).value or ""),
            "source_page": 0,
            "series_name": current["series_name"],
            "episode_or_batch": str(sheet.cell(row, 2).value or "完整作品"),
        }
        current["releases"].append(release)

    for group in groups:
        group["releases"].sort(key=lambda item: (item["published_at"], item["topic_id"]), reverse=True)
        group["latest_at"] = group["releases"][0]["published_at"] if group["releases"] else ""
    return [group for group in groups if group["releases"]]


def release_identity(release: dict) -> tuple[str, str]:
    if release.get("topic_id"):
        return "topic", str(release["topic_id"])
    return "url", str(release.get("detail_url") or release.get("magnet"))


def merge_group_sets(existing_groups: list[dict], new_groups: list[dict]) -> list[dict]:
    releases: dict[tuple[str, str], dict] = {}
    for groups in (existing_groups, new_groups):
        for group in groups:
            for source_release in group["releases"]:
                release = dict(source_release)
                release["series_name"] = group["series_name"]
                releases[release_identity(release)] = release

    grouped: dict[str, list[dict]] = defaultdict(list)
    for release in releases.values():
        grouped[release["series_name"]].append(release)
    merged = []
    for series_name, series_releases in grouped.items():
        series_releases.sort(key=lambda item: (item["published_at"], item["topic_id"]), reverse=True)
        merged.append(
            {"series_name": series_name, "latest_at": series_releases[0]["published_at"], "releases": series_releases}
        )
    merged.sort(key=lambda group: (group["latest_at"], group["series_name"]), reverse=True)
    return merged


def write_xlsx_sheet(
    groups: list[dict], output_path: Path, sheet_name: str, merge_existing: bool = False
) -> tuple[Path, int, int]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as error:
        raise RuntimeError("Excel output requires openpyxl: python3 -m pip install openpyxl") from error

    if not sheet_name or len(sheet_name) > 31 or re.search(r"[\\/*?:\[\]]", sheet_name):
        raise ValueError("Excel sheet name must be 1-31 characters and contain no \\ / * ? : [ ]")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        workbook = load_workbook(output_path)
        template = next(
            (
                candidate
                for candidate in workbook.worksheets
                if [candidate.cell(1, column).value for column in range(1, 10)]
                == ["片名", "集数/批次", "发布日期", "大小", "分类", "发布者", "发布标题", "详情页", "磁链"]
            ),
            None,
        )
    else:
        workbook = Workbook()
        template = None

    existing = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
    new_release_keys: set[tuple[str, str]] = set()
    if merge_existing and existing is not None:
        existing_groups = read_xlsx_groups(existing)
        existing_release_keys = {
            release_identity(release)
            for group in existing_groups
            for release in group["releases"]
        }
        new_release_keys = {
            release_identity(release)
            for group in groups
            for release in group["releases"]
        } - existing_release_keys
        groups = merge_group_sets(existing_groups, groups)

    template_group_row = 2
    template_release_row = 3
    if template is not None:
        for candidate_row in range(2, template.max_row + 1):
            fill_rgb = str(template.cell(candidate_row, 1).fill.fgColor.rgb or "").upper()
            if (
                template.row_dimensions[candidate_row].collapsed
                and template.cell(candidate_row, 1).value
                and not fill_rgb.endswith(NEW_ITEM_FILL_COLOR)
            ):
                template_group_row = candidate_row
                break
        for candidate_row in range(2, template.max_row + 1):
            fill_rgb = str(template.cell(candidate_row, 2).fill.fgColor.rgb or "").upper()
            if template.row_dimensions[candidate_row].hidden and not fill_rgb.endswith(NEW_ITEM_FILL_COLOR):
                template_release_row = candidate_row
                break
    index = workbook.index(existing) if existing is not None else len(workbook.worksheets)
    temporary_title = "__dmhy_new__"
    while temporary_title in workbook.sheetnames:
        temporary_title += "_"
    sheet = workbook.create_sheet(temporary_title, index=index)
    if existing is not None:
        workbook.remove(existing)
    sheet.title = sheet_name

    default_sheet = workbook["Sheet"] if "Sheet" in workbook.sheetnames else None
    if default_sheet is not None and default_sheet.max_row == 1 and default_sheet["A1"].value is None:
        workbook.remove(default_sheet)

    headers = ["片名", "集数/批次", "发布日期", "大小", "分类", "发布者", "发布标题", "详情页", "磁链"]
    widths = [66, 14, 18, 11, 10, 20, 82, 11, 68]
    new_item_fill = PatternFill("solid", fgColor=NEW_ITEM_FILL_COLOR)
    normal_group_fill = PatternFill("solid", fgColor="DDEBE4")
    normal_release_fill = PatternFill()
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.outlinePr.summaryBelow = False
    sheet.sheet_properties.outlinePr.applyStyles = True
    sheet.sheet_format.defaultRowHeight = 15

    for column, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = sheet.cell(1, column, header)
        sheet.column_dimensions[cell.column_letter].width = width
        if template is not None:
            cell._style = copy(template.cell(1, column)._style)
        else:
            cell.fill = PatternFill("solid", fgColor="1F513A")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = template.row_dimensions[1].height if template is not None else 26

    row_number = 2
    for group in groups:
        group_has_new_release = any(
            release_identity(release) in new_release_keys for release in group["releases"]
        )
        group_row = row_number
        sheet.cell(group_row, 1, group["series_name"])
        for column in range(1, 10):
            cell = sheet.cell(group_row, column)
            if template is not None:
                cell._style = copy(template.cell(template_group_row, column)._style)
            else:
                cell.font = Font(bold=True, color="173E2F")
                cell.alignment = Alignment(vertical="center")
            cell.fill = new_item_fill if group_has_new_release else normal_group_fill
        sheet.row_dimensions[group_row].height = (
            template.row_dimensions[template_group_row].height if template is not None else 25
        )
        sheet.row_dimensions[group_row].collapsed = True
        row_number += 1

        for release in group["releases"]:
            release_is_new = release_identity(release) in new_release_keys
            values = [
                None,
                release["episode_or_batch"],
                release["published_at"],
                release["size"],
                release["category"],
                release["publisher"],
                release["title"],
                "详情页",
                release["magnet"],
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row_number, column, value)
                if template is not None:
                    cell._style = copy(template.cell(template_release_row, column)._style)
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=column in {7, 9})
                cell.fill = new_item_fill if release_is_new else normal_release_fill
            sheet.cell(row_number, 8).hyperlink = release["detail_url"]
            sheet.cell(row_number, 9).hyperlink = release["magnet"]
            if template is None:
                for column in (8, 9):
                    sheet.cell(row_number, column).font = Font(color="0563C1", underline="single")
            sheet.row_dimensions[row_number].height = (
                template.row_dimensions[template_release_row].height if template is not None else 31
            )
            sheet.row_dimensions[row_number].hidden = True
            sheet.row_dimensions[row_number].outlineLevel = 1
            row_number += 1

    sheet.auto_filter.ref = f"A1:I{sheet.max_row}"
    workbook.active = workbook.index(sheet)

    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f"{output_path.stem}.", suffix=".tmp.xlsx", dir=output_path.parent
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        checked_workbook = load_workbook(temporary_path, data_only=False)
        checked_sheet = checked_workbook[sheet_name]
        release_count = sum(len(group["releases"]) for group in groups)
        hidden_count = sum(
            bool(checked_sheet.row_dimensions[row].hidden) for row in range(2, checked_sheet.max_row + 1)
        )
        valid_magnets = sum(
            bool(checked_sheet.cell(row, 9).hyperlink)
            and checked_sheet.cell(row, 9).hyperlink.target == checked_sheet.cell(row, 9).value
            for row in range(2, checked_sheet.max_row + 1)
        )
        highlighted_release_count = sum(
            bool(checked_sheet.row_dimensions[row].hidden)
            and str(checked_sheet.cell(row, 2).fill.fgColor.rgb or "").upper().endswith(NEW_ITEM_FILL_COLOR)
            for row in range(2, checked_sheet.max_row + 1)
        )
        highlighted_series_count = sum(
            bool(checked_sheet.row_dimensions[row].collapsed)
            and str(checked_sheet.cell(row, 1).fill.fgColor.rgb or "").upper().endswith(NEW_ITEM_FILL_COLOR)
            for row in range(2, checked_sheet.max_row + 1)
        )
        expected_series_count = sum(
            any(release_identity(release) in new_release_keys for release in group["releases"])
            for group in groups
        )
        checked_workbook.close()
        if checked_sheet.max_row != 1 + len(groups) + release_count:
            raise RuntimeError("Excel row-count validation failed")
        if hidden_count != release_count or valid_magnets != release_count:
            raise RuntimeError("Excel outline or magnet-link validation failed")
        if highlighted_release_count != len(new_release_keys) or highlighted_series_count != expected_series_count:
            raise RuntimeError("Excel new-release highlighting validation failed")
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()
    return output_path, len(new_release_keys), expected_series_count


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", nargs="?", help="DMHY team/user URL or numeric team ID")
    parser.add_argument("--from-json", type=Path, help="organize an existing scraper JSON without network access")
    parser.add_argument("--last-page", type=int, help="skip automatic last-page discovery")
    parser.add_argument("--since-days", type=int, help="scrape only releases from the latest N days")
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--output-dir", type=Path, default=Path.cwd())
    parser.add_argument("--raw-name", "--name", dest="raw_name", help="raw CSV/JSON basename")
    parser.add_argument("--grouped-name", help="grouped CSV/HTML basename")
    parser.add_argument("--group-name", help="override the detected group/publisher name")
    parser.add_argument("--title-contains", help="keep only releases whose title contains this exact text")
    parser.add_argument("--xlsx", type=Path, help="create or update an Excel workbook with a collapsible sheet")
    parser.add_argument("--sheet-name", help="Excel sheet name; defaults to the group name")
    parser.add_argument("--merge-xlsx", action="store_true", help="merge scraped releases into an existing sheet")
    parser.add_argument("--raw-only", action="store_true", help="skip grouped CSV and HTML")
    args = parser.parse_args()

    if (args.source is None) == (args.from_json is None):
        parser.error("provide either a team source or --from-json, but not both")
    if args.from_json is not None and args.raw_only and args.xlsx is None:
        parser.error("--raw-only with --from-json requires --xlsx")
    if args.sheet_name is not None and args.xlsx is None:
        parser.error("--sheet-name requires --xlsx")
    if args.last_page is not None and args.last_page < 1:
        parser.error("--last-page must be positive")
    if args.since_days is not None and args.since_days < 1:
        parser.error("--since-days must be positive")
    if args.merge_xlsx and args.xlsx is None:
        parser.error("--merge-xlsx requires --xlsx")
    if not 1 <= args.workers <= 8:
        parser.error("--workers must be between 1 and 8")

    try:
        args.output_dir.mkdir(parents=True, exist_ok=True)
        output_paths: list[Path] = []
        pages: int | None = None
        if args.from_json is not None:
            team_name, rows = load_export(args.from_json)
        else:
            _, list_url = normalize_source(args.source)
            rows, pages, team_name = scrape(list_url, args.last_page, args.workers, args.since_days)

        if args.group_name:
            team_name = args.group_name
        if args.title_contains:
            rows = [row for row in rows if args.title_contains in row["title"]]
            if not rows:
                raise RuntimeError(f"no releases matched --title-contains {args.title_contains!r}")
        validate_rows(rows)

        if args.from_json is None:
            raw_basename = safe_basename(args.raw_name or f"{team_name}_topics")
            raw_csv, raw_json = write_raw_outputs(
                rows, args.output_dir, raw_basename, list_url, pages, team_name
            )
            output_paths.extend((raw_csv, raw_json))

        group_count: int | None = None
        if not args.raw_only:
            grouped_basename = safe_basename(args.grouped_name or f"{team_name}_by_series")
            grouped_csv, grouped_html, group_count = write_grouped_outputs(
                team_name, rows, args.output_dir, grouped_basename
            )
            output_paths.extend((grouped_csv, grouped_html))
        highlighted_release_count: int | None = None
        highlighted_series_count: int | None = None
        if args.xlsx is not None:
            groups = group_items(rows)
            xlsx_path, highlighted_release_count, highlighted_series_count = write_xlsx_sheet(
                groups, args.xlsx, args.sheet_name or team_name, merge_existing=args.merge_xlsx
            )
            output_paths.append(xlsx_path)
            group_count = len(groups)
    except (OSError, ValueError, RuntimeError, subprocess.SubprocessError) as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1

    if pages is not None:
        print(f"Processed {len(rows)} unique topics from {team_name} across {pages} pages")
    else:
        print(f"Processed {len(rows)} unique topics from existing JSON for {team_name}")
    if group_count is not None:
        print(f"Grouped into {group_count} series")
    if args.merge_xlsx and highlighted_release_count is not None:
        print(
            f"Highlighted {highlighted_release_count} new releases "
            f"across {highlighted_series_count} series"
        )
    print(f"Date range: {rows[-1]['published_at']} to {rows[0]['published_at']}")
    for path in output_paths:
        print(path.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
