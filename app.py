#!/usr/bin/env python3
"""Local browser UI for incremental DMHY workbook updates."""

from __future__ import annotations

import argparse
import csv
import email.utils
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import time
import webbrowser
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlencode, urlparse
from zoneinfo import ZoneInfo

import dmhy_scraper as scraper
from openpyxl import load_workbook


APP_DIR = Path(__file__).resolve().parent
STATIC_DIR = APP_DIR / "static"
DATA_DIR = APP_DIR.parent
SCRAPER = APP_DIR / "dmhy_scraper.py"
VERSION_FILE = APP_DIR / "updates.json"
WORKBOOK_RE = re.compile(r"^DMHY更新表_(\d{4}-\d{2}-\d{2})\.xlsx$")
TOPIC_ID_RE = re.compile(r"/topics/view/(\d+)")
NEW_FILL_COLOR = "C6EFCE"
SHANGHAI = ZoneInfo("Asia/Shanghai")
FALLBACK_DMHY_IP = "104.25.61.106"

GROUPS = {
    "lolihouse": {
        "label": "LoliHouse",
        "source": "https://share.dmhy.org/topics/list/team_id/657",
        "rss": "https://share.dmhy.org/topics/rss/team_id/657",
        "sheet": "LoliHouse",
        "output_dir": DATA_DIR,
        "raw_name": "lolihouse_topics",
        "grouped_name": "lolihouse_by_series",
        "csv": DATA_DIR / "lolihouse_topics.csv",
        "extra_args": [],
    },
    "7acg": {
        "label": "7³ACG",
        "source": "https://share.dmhy.org/topics/list/user_id/759200",
        "rss": "https://share.dmhy.org/topics/rss/rss.xml",
        "sheet": "7³ACG",
        "output_dir": DATA_DIR / "7acg",
        "raw_name": "7acg_topics",
        "grouped_name": "7acg_by_series",
        "csv": DATA_DIR / "7acg" / "7acg_topics.csv",
        "extra_args": ["--title-contains", "[7³ACG]", "--group-name", "7³ACG"],
    },
}

STATE_LOCK = threading.Lock()
STATE = {
    "status": "idle",
    "message": "准备就绪",
    "progress": 0,
    "days": 7,
    "groups": [],
    "results": [],
    "logs": [],
    "workbook": None,
    "started_at": None,
    "finished_at": None,
    "error": None,
}
RECENT_CACHE: dict = {
    "workbook": None,
    "mtime": 0.0,
    "releases": [],
    "ready": False,
}
RECENT_CACHE_LOCK = threading.Lock()
DEMO_MODE = False


def now_iso() -> str:
    return datetime.now(SHANGHAI).isoformat(timespec="seconds")


def update_state(**changes) -> None:
    with STATE_LOCK:
        STATE.update(changes)


def append_log(line: str) -> None:
    clean = line.strip()
    if not clean:
        return
    with STATE_LOCK:
        STATE["logs"] = (STATE["logs"] + [clean])[-80:]


def load_versions() -> dict:
    if VERSION_FILE.exists():
        try:
            data = json.loads(VERSION_FILE.read_text(encoding="utf-8"))
            if "last_update" not in data:
                data["last_update"] = None
                save_versions(data)
            if "last_update_by_group" not in data:
                data["last_update_by_group"] = {}
                save_versions(data)
            return data
        except (OSError, ValueError):
            pass
    initial = {
        "current": "0.0.0.1",
        "last_update": None,
        "entries": [
            {
                "version": "0.0.0.1",
                "time": datetime.now(SHANGHAI).strftime("%Y-%m-%d %H:%M"),
                "summary": "初始版本",
                "details": [],
            }
        ],
    }
    save_versions(initial)
    return initial


def save_versions(data: dict) -> None:
    VERSION_FILE.write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def set_last_update(group_id: str | None = None) -> None:
    """Record a successful update. With a group id, write only that group's
    timestamp; the shared `last_update` stays as the last fully-successful run
    so groups without their own entry keep the full fallback window."""
    data = load_versions()
    stamp = now_iso()
    if group_id:
        data.setdefault("last_update_by_group", {})[group_id] = stamp
    else:
        data["last_update"] = stamp
    save_versions(data)


def compute_since_days(group_id: str, fallback_days: int) -> int:
    """Incrementally fetch only releases published since that group's last
    successful update; falls back to the shared timestamp for older files."""
    data = load_versions()
    last = data.get("last_update_by_group", {}).get(group_id) or data.get("last_update")
    if not last:
        return fallback_days
    try:
        last_dt = datetime.fromisoformat(str(last))
        if last_dt.tzinfo is None:
            last_dt = last_dt.replace(tzinfo=SHANGHAI)
        now = datetime.now(SHANGHAI)
        if now <= last_dt:
            return 1
        return max(1, (now - last_dt).days + 1)
    except ValueError:
        return fallback_days


def bump_version(summary: str, details: list[str]) -> str:
    data = load_versions()
    parts = data["current"].split(".")
    parts[-1] = str(int(parts[-1]) + 1)
    new_version = ".".join(parts)
    data["current"] = new_version
    data["entries"].append(
        {
            "version": new_version,
            "time": datetime.now(SHANGHAI).strftime("%Y-%m-%d %H:%M"),
            "summary": summary,
            "details": details,
        }
    )
    save_versions(data)
    return new_version


def state_snapshot() -> dict:
    with STATE_LOCK:
        return json.loads(json.dumps(STATE, ensure_ascii=False))


def latest_workbook() -> Path | None:
    candidates: list[tuple[str, Path]] = []
    for path in DATA_DIR.glob("DMHY更新表_*.xlsx"):
        match = WORKBOOK_RE.match(path.name)
        if match:
            candidates.append((match.group(1), path))
    return max(candidates, default=("", None), key=lambda item: item[0])[1]


def is_new_fill(cell) -> bool:
    return str(cell.fill.fgColor.rgb or "").upper().endswith(NEW_FILL_COLOR)


def highlighted_releases(workbook_path: Path, group_ids: list[str]) -> list[dict]:
    if not workbook_path.exists():
        return []
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    results: list[dict] = []
    try:
        for group_id in group_ids:
            config = GROUPS[group_id]
            if config["sheet"] not in workbook.sheetnames:
                continue
            sheet = workbook[config["sheet"]]
            series_name = ""
            for row in range(2, sheet.max_row + 1):
                dimension = sheet.row_dimensions[row]
                if dimension.collapsed and sheet.cell(row, 1).value:
                    series_name = str(sheet.cell(row, 1).value)
                    continue
                if not dimension.hidden or not is_new_fill(sheet.cell(row, 2)):
                    continue
                detail_cell = sheet.cell(row, 8)
                magnet_cell = sheet.cell(row, 9)
                detail_url = detail_cell.hyperlink.target if detail_cell.hyperlink else ""
                magnet = magnet_cell.hyperlink.target if magnet_cell.hyperlink else str(magnet_cell.value or "")
                topic_match = TOPIC_ID_RE.search(detail_url)
                results.append(
                    {
                        "group_id": group_id,
                        "group": config["label"],
                        "series": series_name,
                        "episode": str(sheet.cell(row, 2).value or "完整作品"),
                        "published_at": str(sheet.cell(row, 3).value or ""),
                        "size": str(sheet.cell(row, 4).value or ""),
                        "category": str(sheet.cell(row, 5).value or ""),
                        "publisher": str(sheet.cell(row, 6).value or ""),
                        "title": str(sheet.cell(row, 7).value or ""),
                        "detail_url": detail_url,
                        "magnet": magnet,
                        "topic_id": int(topic_match.group(1)) if topic_match else 0,
                    }
                )
    finally:
        workbook.close()
    return sorted(results, key=lambda item: (item["published_at"], item["topic_id"]), reverse=True)


CSV_FIELDS = [
    "topic_id",
    "published_at",
    "category",
    "title",
    "detail_url",
    "magnet",
    "size",
    "seeders",
    "downloads",
    "completed",
    "publisher",
    "source_page",
]


def read_csv_rows(path: Path) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv_rows(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def merge_csv_rows(rows: list[dict], path: Path, existing: list[dict] | None = None) -> int:
    """Merge new rows into the CSV by topic ID and return the new count."""
    if existing is None:
        existing = read_csv_rows(path)
    by_id: dict[int, dict] = {}
    for row in existing:
        try:
            by_id[int(row["topic_id"])] = row
        except (KeyError, ValueError):
            continue
    new_count = 0
    for row in rows:
        try:
            topic_id = int(row["topic_id"])
        except (KeyError, ValueError):
            continue
        if topic_id not in by_id:
            new_count += 1
        by_id[topic_id] = row
    merged = sorted(
        by_id.values(),
        key=lambda row: (str(row.get("published_at", "")), int(row.get("topic_id") or 0)),
        reverse=True,
    )
    write_csv_rows(path, merged)
    return new_count


def parse_published(raw: str):
    value = str(raw or "").strip()[:19]
    for fmt in ("%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(value, fmt).replace(tzinfo=SHANGHAI)
        except ValueError:
            continue
    return None


def parse_title(title: str, group_id: str) -> tuple[str, str]:
    text = str(title or "").strip()
    # 去掉开头的字幕组标签（兼容半角 [组] 与全角 【组】）
    match = re.match(r"^[【\[]([^】\]]+)[】\]]\s*", text)
    if match:
        text = text[match.end():]
    if group_id == "7acg":
        parts = re.split(r"\s*\|\s*", text, maxsplit=1)
        if len(parts) == 1:
            series = re.split(r"\s+\[", text, maxsplit=1)[0].strip()
            return series, "完整作品"
    else:
        parts = re.split(r"\s+-\s+", text, maxsplit=1)
        if len(parts) == 1:
            # 旧式全角/半角括号标题：【字幕组】【番名】【集数】【技术信息】
            bracket_parts = re.findall(r"[【\[]([^】\]]*)[】\]]", text)
            if bracket_parts:
                episode = None
                series_parts = bracket_parts
                if re.fullmatch(r"(?:第\s*)?\d+(?:\s*[話话])?", bracket_parts[-1].strip()):
                    episode, series_parts = bracket_parts[-1].strip(), bracket_parts[:-1]
                elif len(bracket_parts) >= 3 and re.fullmatch(
                    r"(?:第\s*)?\d+(?:\s*[話话])?", bracket_parts[-2].strip()
                ):
                    episode, series_parts = bracket_parts[-2].strip(), bracket_parts[:-2]
                series = " / ".join(p.strip() for p in series_parts if p.strip())
                if series:
                    return series, episode or "完整作品"
            series = re.split(r"\s+\[", text, maxsplit=1)[0].strip()
            return series, "完整作品"
    series = parts[0].strip() if parts else text
    episode = re.split(r"\s+\[", parts[1].strip(), maxsplit=1)[0].strip() if len(parts) > 1 else "完整作品"
    return series, episode


def read_all_releases(_workbook_path: Path | None = None) -> list[dict]:
    """Return every release from the merged CSVs, newest first."""
    results: list[dict] = []
    for group_id, config in GROUPS.items():
        for row in read_csv_rows(config["csv"]):
            published_raw = str(row.get("published_at", "")).replace("/", "-")
            published_dt = parse_published(str(row.get("published_at", "")))
            series, episode = parse_title(row.get("title", ""), group_id)
            try:
                topic_id = int(row["topic_id"])
            except (KeyError, ValueError):
                topic_id = 0
            results.append(
                {
                    "group_id": group_id,
                    "group": config["label"],
                    "series": series,
                    "episode": episode,
                    "published_at": published_raw,
                    "size": str(row.get("size", "") or ""),
                    "category": str(row.get("category", "") or ""),
                    "publisher": str(row.get("publisher", "") or ""),
                    "title": str(row.get("title", "") or ""),
                    "detail_url": str(row.get("detail_url", "") or ""),
                    "magnet": str(row.get("magnet", "") or ""),
                    "topic_id": topic_id,
                    "published": published_dt,
                }
            )
    return sorted(
        results,
        key=lambda item: (item["published_at"], item["topic_id"]),
        reverse=True,
    )


def cached_releases(group_ids: list[str] | None = None) -> list[dict]:
    """Every release from the merged CSVs currently loaded in the cache."""
    with RECENT_CACHE_LOCK:
        releases = list(RECENT_CACHE["releases"])
    if group_ids:
        releases = [item for item in releases if item.get("group_id") in group_ids]
    return releases


def search_releases(needle: str, group_ids: list[str] | None = None) -> list[dict]:
    """Search the already-loaded releases by series/episode/title/group name."""
    needle = needle.casefold()
    results = []
    for item in cached_releases(group_ids):
        if (
            needle in (item.get("series") or "").casefold()
            or needle in (item.get("episode") or "").casefold()
            or needle in (item.get("title") or "").casefold()
            or needle in (item.get("group") or "").casefold()
        ):
            results.append(item)
    return results


def recent_releases(days: int, group_ids: list[str] | None = None) -> list[dict]:
    now = datetime.now(SHANGHAI)
    cutoff = now - timedelta(days=days)
    return [
        item
        for item in cached_releases(group_ids)
        if item.get("published") and item["published"] >= cutoff
    ]


def refresh_recent_cache() -> dict:
    signature = "|".join(
        f"{config['csv'].name}:{config['csv'].stat().st_mtime if config['csv'].exists() else 0}"
        for config in GROUPS.values()
    )
    acquired = RECENT_CACHE_LOCK.acquire(timeout=0.1)
    if not acquired:
        return {
            "workbook": latest_workbook(),
            "releases": [],
            "ready": False,
        }
    try:
        if RECENT_CACHE["mtime"] == signature:
            return dict(RECENT_CACHE)
        releases = read_all_releases()
        RECENT_CACHE.update(
            {"mtime": signature, "releases": releases, "ready": True}
        )
        return dict(RECENT_CACHE)
    finally:
        RECENT_CACHE_LOCK.release()


def prepare_workbook() -> tuple[Path, Path]:
    today = datetime.now(SHANGHAI).strftime("%Y-%m-%d")
    target = DATA_DIR / f"DMHY更新表_{today}.xlsx"
    source = target if target.exists() else latest_workbook()
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f"DMHY更新表_{today}.", suffix=".updating.xlsx", dir=DATA_DIR
    )
    os.close(descriptor)
    working = Path(temporary_name)
    if source is not None:
        shutil.copy2(source, working)
    else:
        working.unlink()
    return target, working


def command_for(group_id: str, days: int) -> list[str]:
    config = GROUPS[group_id]
    config["output_dir"].mkdir(parents=True, exist_ok=True)
    return [
        sys.executable,
        str(SCRAPER),
        config["source"],
        "--since-days",
        str(days),
        "--output-dir",
        str(config["output_dir"]),
        "--raw-name",
        config["raw_name"],
        "--grouped-name",
        config["grouped_name"],
        *config["extra_args"],
    ]


def run_command(command: list[str], environment: dict[str, str]) -> tuple[int, str]:
    process = subprocess.Popen(
        command,
        cwd=APP_DIR,
        env=environment,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        bufsize=1,
    )
    lines: list[str] = []
    assert process.stdout is not None
    for line in process.stdout:
        lines.append(line)
        append_log(line)
    return process.wait(), "".join(lines)


def friendly_update_error(output: str) -> str:
    """Turn raw scraper output into a clear, actionable error message."""
    lowered = output.lower()
    if "curl: (56)" in output or "http 500" in lowered:
        if "searchd" in lowered or "服务器遇到错误" in output:
            return "DMHY 服务器暂时故障（其搜索服务不可用，页面返回 500），请稍后重试"
        return "DMHY 服务器暂时故障（页面返回 500），请稍后重试"
    return next((line for line in reversed(output.splitlines()) if line.strip()), "未知错误")


def update_group(group_id: str, days: int) -> None:
    config = GROUPS[group_id]
    command = command_for(group_id, days)
    environment = os.environ.copy()
    environment["PYTHONUNBUFFERED"] = "1"
    if Path("/etc/ssl/cert.pem").exists():
        environment["SSL_CERT_FILE"] = "/etc/ssl/cert.pem"
    # Local DNS may be polluted; the verified Cloudflare IP is the fast,
    # reliable route, so start with it and keep the direct route as backup.
    environment.setdefault("DMHY_RESOLVE_IP", FALLBACK_DMHY_IP)

    code, output = run_command(command, environment)
    if code == 0:
        return
    connection_failure = any(
        token in output for token in ("curl: (6)", "curl: (7)", "curl: (28)")
    )
    if connection_failure:
        append_log(f"{config['label']} 备用线路连接失败，正在切换直连重试")
        update_state(message=f"{config['label']} 正在重试…")
        environment.pop("DMHY_RESOLVE_IP", None)
        code, output = run_command(command, environment)
    if code != 0:
        raise RuntimeError(f"{config['label']} 更新失败：{friendly_update_error(output)}")


def group_title_filter(config: dict) -> str:
    """Extract the --title-contains value from a group's scraper args, if any."""
    args = config.get("extra_args") or []
    for index, arg in enumerate(args):
        if arg == "--title-contains" and index + 1 < len(args):
            return args[index + 1]
    return ""


def fetch_url(url: str, environment: dict) -> str:
    """Fetch a URL with curl, honoring the DMHY_RESOLVE_IP override."""
    command = [
        "curl",
        "-L",
        "--compressed",
        "--fail",
        "--silent",
        "--show-error",
        "--connect-timeout",
        "15",
        "--max-time",
        "60",
        "-A",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    ]
    resolve_ip = environment.get("DMHY_RESOLVE_IP", "").strip()
    if resolve_ip:
        command += ["--resolve", f"share.dmhy.org:443:{resolve_ip}"]
    command.append(url)
    process = subprocess.run(
        command,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=90,
    )
    if process.returncode != 0:
        detail = (process.stderr or process.stdout or "").strip().splitlines()
        raise RuntimeError(detail[-1] if detail else f"curl exit {process.returncode}")
    return process.stdout


def fetch_url_with_fallback(url: str, environment: dict) -> str | None:
    """Fetch a URL, preferring the verified fallback IP first (local DNS may be
    polluted and make the direct route hang), then the direct route."""
    routes: list[dict] = []
    base = dict(environment)
    ip_route = dict(environment)
    ip_route["DMHY_RESOLVE_IP"] = FALLBACK_DMHY_IP
    if environment.get("DMHY_RESOLVE_IP"):
        routes = [base, ip_route]
    else:
        routes = [ip_route, base]
    for route in routes:
        try:
            return fetch_url(url, route)
        except Exception:
            continue
    return None


def parse_rss_items(feed: str) -> list[dict]:
    """Parse <item> blocks from a DMHY RSS feed (title/link/pubDate)."""
    items: list[dict] = []
    for block in re.findall(r"<item>(.*?)</item>", feed, flags=re.S):
        title = re.search(r"<title><!\[CDATA\[(.*?)\]\]></title>", block, flags=re.S)
        link = re.search(r"<link>(.*?)</link>", block, flags=re.S)
        pub = re.search(r"<pubDate>(.*?)</pubDate>", block, flags=re.S)
        if not title or not link or not pub:
            continue
        items.append(
            {
                "title": title.group(1).strip(),
                "link": link.group(1).strip(),
                "pub_date": pub.group(1).strip(),
            }
        )
    return items


def parse_topic_detail(detail_url: str, page: str) -> dict | None:
    """Extract a full CSV row from a topic detail page (fields the RSS omits)."""
    magnet_match = re.search(r'<a id="magnet2" href="(magnet:[^"]+)"', page) or re.search(
        r'data-magnet="(magnet:[^"]+)"', page
    )
    topic_match = TOPIC_ID_RE.search(detail_url)
    if not magnet_match or not topic_match:
        return None
    category = "未分类"
    category_segment = re.search(r"所屬分類: <span>(.*?)<span></li>", page, flags=re.S)
    if category_segment:
        matches = re.findall(r">([^<>]+)</", category_segment.group(1))
        if matches:
            category = matches[-1].strip() or category
    size_match = re.search(r"文件大小: <span>([^<]+)</span>", page)
    time_match = re.search(r"發佈時間: <span>([^<]+)</span>", page)
    publisher_match = re.search(r"發佈人：<a href=\"[^\"]*\">([^<]+)</a>", page)
    return {
        "topic_id": int(topic_match.group(1)),
        "published_at": (time_match.group(1).replace("/", "-") if time_match else ""),
        "category": category,
        "title": "",
        "detail_url": detail_url,
        "magnet": magnet_match.group(1).split("&", 1)[0],
        "size": size_match.group(1).strip() if size_match else "-",
        "seeders": "-",
        "downloads": "-",
        "completed": "-",
        "publisher": publisher_match.group(1).strip() if publisher_match else "",
        "source_page": 0,
    }


def rss_fallback_rows(config: dict, since_days: int, environment: dict) -> tuple[bool, list[dict]]:
    """Fallback source when the listing page fails (e.g. DMHY search service
    down): read the group's RSS feed, filter matching items within the window,
    then fetch each detail page for the fields the feed omits.

    Returns (ok, rows); ok=False means the feed itself could not be fetched."""
    rss_url = config.get("rss")
    if not rss_url:
        return False, []
    feed = fetch_url_with_fallback(rss_url, environment)
    if feed is None:
        return False, []
    cutoff = datetime.now(SHANGHAI) - timedelta(days=since_days)
    title_filter = group_title_filter(config)
    matches: list[tuple[dict, datetime]] = []
    for item in parse_rss_items(feed):
        if title_filter and title_filter.casefold() not in item["title"].casefold():
            continue
        try:
            published = email.utils.parsedate_to_datetime(item["pub_date"])
        except (TypeError, ValueError):
            continue
        if published is None or published < cutoff:
            continue
        if not TOPIC_ID_RE.search(item["link"]):
            continue
        matches.append((item, published))

    rows = complete_rss_matches(matches, environment)
    return True, rows


def complete_rss_matches(
    matches: list[tuple[dict, datetime]], environment: dict, workers: int = 4
) -> list[dict]:
    """Fetch each RSS item's detail page in parallel and build full rows."""

    def fetch_one(match: tuple[dict, datetime]) -> dict | None:
        item, published = match
        detail_url = item["link"].replace("http://", "https://")
        page = fetch_url_with_fallback(detail_url, environment)
        if page is None:
            return None
        row = parse_topic_detail(detail_url, page)
        if row is None:
            return None
        row["title"] = item["title"]
        row["published_at"] = published.astimezone(SHANGHAI).strftime("%Y-%m-%d %H:%M")
        return row

    with ThreadPoolExecutor(max_workers=workers) as pool:
        return [row for row in pool.map(fetch_one, matches) if row is not None]


SITE_SEARCH_URL = "https://share.dmhy.org/topics/list"
SITE_RSS_URL = "https://share.dmhy.org/topics/rss/rss.xml"
SITE_SEARCH_CACHE_TTL = 120.0
SITE_SEARCH_RSS_LIMIT = 60
site_search_cache: dict[str, dict] = {}
site_search_cache_lock = threading.Lock()


def site_group_id(title: str) -> str:
    """Map a DMHY title to the app's group ids (for filter compatibility)."""
    if title.startswith("[LoliHouse]"):
        return "lolihouse"
    if title.startswith("[7³ACG]") or title.startswith("[7³ACG]"):
        return "7acg"
    return "site"


def site_group_label(title: str) -> str:
    match = re.match(r"^[【\[]([^】\]]+)[】\]]", title.strip())
    return match.group(1) if match else "DMHY"


def release_from_topic_row(row: dict) -> dict:
    """Shape a scraper/RSS topic row like the /api/recent release items."""
    title = str(row.get("title", "") or "")
    group_id = site_group_id(title)
    series, episode = parse_title(title, "7acg" if group_id == "7acg" else "lolihouse")
    try:
        topic_id = int(row.get("topic_id") or 0)
    except (TypeError, ValueError):
        topic_id = 0
    return {
        "group_id": group_id,
        "group": site_group_label(title),
        "series": series,
        "episode": episode,
        "published_at": str(row.get("published_at", "") or "").replace("/", "-"),
        "size": str(row.get("size", "") or "-"),
        "category": str(row.get("category", "") or ""),
        "publisher": str(row.get("publisher", "") or ""),
        "title": title,
        "detail_url": str(row.get("detail_url", "") or ""),
        "magnet": str(row.get("magnet", "") or ""),
        "topic_id": topic_id,
    }


def dmhy_site_search(query: str, group_id: str, environment: dict) -> tuple[str, list[dict]]:
    """Search DMHY site-wide by keyword. Returns (source, release rows).

    Prefers DMHY's own keyword search (complete results with magnets); when
    that service is unavailable (searchd outages return HTTP 500) it falls
    back to filtering the site-wide RSS feed, which only covers the most
    recent topics."""
    params = {"keyword": query}
    if group_id == "lolihouse":
        params["team_id"] = "657"
    elif group_id == "7acg":
        params["user_id"] = "759200"
    page = fetch_url_with_fallback(SITE_SEARCH_URL + "?" + urlencode(params), environment)
    if page is not None:
        try:
            parser = scraper.parse_page(1, page)
            return "site", [release_from_topic_row(row) for row in parser.rows]
        except Exception:
            pass
    feed = fetch_url_with_fallback(SITE_RSS_URL, environment)
    if feed is None:
        raise RuntimeError("无法连接 DMHY（搜索与 RSS 均失败），请稍后重试")
    needle = query.casefold()
    matches: list[tuple[dict, datetime]] = []
    for item in parse_rss_items(feed):
        if needle not in item["title"].casefold():
            continue
        if group_id in GROUPS and site_group_id(item["title"]) != group_id:
            continue
        try:
            published = email.utils.parsedate_to_datetime(item["pub_date"])
        except (TypeError, ValueError):
            continue
        if published is None or not TOPIC_ID_RE.search(item["link"]):
            continue
        matches.append((item, published))
    rows = complete_rss_matches(matches[:SITE_SEARCH_RSS_LIMIT], environment, workers=6)
    return "rss", [release_from_topic_row(row) for row in rows]


def site_search_cached(query: str, group_id: str) -> dict:
    """Serve a site search through a short-lived cache to spare DMHY."""
    key = f"{group_id}|{query.strip().casefold()}"
    with site_search_cache_lock:
        cached = site_search_cache.get(key)
        if cached and time.time() - cached["time"] < SITE_SEARCH_CACHE_TTL:
            return cached["payload"]
    environment = os.environ.copy()
    source, rows = dmhy_site_search(query.strip(), group_id, environment)
    payload = {"q": query.strip(), "source": source, "total": len(rows), "results": rows}
    with site_search_cache_lock:
        if len(site_search_cache) > 40:
            ordered = sorted(site_search_cache.items(), key=lambda kv: kv[1]["time"])
            for stale_key, _ in ordered[: len(site_search_cache) - 20]:
                site_search_cache.pop(stale_key, None)
        site_search_cache[key] = {"time": time.time(), "payload": payload}
    return payload


def generate_excel() -> Path:
    """Generate today's workbook on demand from the latest JSON data."""
    working: Path | None = None
    try:
        target, working = prepare_workbook()
        for group_id, config in GROUPS.items():
            json_path = config["output_dir"] / f"{config['raw_name']}.json"
            if not json_path.exists():
                continue
            command = [
                sys.executable,
                str(SCRAPER),
                "--from-json",
                str(json_path),
                "--raw-only",
                "--merge-xlsx",
                "--xlsx",
                str(working),
                "--sheet-name",
                config["sheet"],
            ]
            environment = os.environ.copy()
            environment["PYTHONUNBUFFERED"] = "1"
            run_command(command, environment)
        if working.exists():
            os.replace(working, target)
            working = None
            append_log(f"Excel 表格已生成：{target.name}")
        return target
    except Exception as error:
        append_log(f"Excel 生成失败：{error}")
        raise
    finally:
        if working is not None and working.exists():
            working.unlink()


def run_demo_update(days: int, group_ids: list[str]) -> None:
    time.sleep(0.35)
    update_state(progress=55, message="正在整理新增记录…")
    time.sleep(0.35)
    releases = read_all_releases()
    results = [item for item in releases if item["group_id"] in group_ids][: days * 5]
    update_state(
        status="success",
        message=f"更新完成，新增 {len(results)} 条",
        progress=100,
        results=results,
        workbook=str(latest_workbook()) if latest_workbook() else None,
        finished_at=now_iso(),
    )


def update_worker(days: int, group_ids: list[str]) -> None:
    if DEMO_MODE:
        run_demo_update(days, group_ids)
        return
    succeeded: list[str] = []
    fallback_used: list[str] = []
    failed: list[str] = []
    total_new = 0
    for index, group_id in enumerate(group_ids):
        config = GROUPS[group_id]
        progress = 8 + round(index / len(group_ids) * 82)
        update_state(progress=progress, message=f"正在更新 {config['label']}…")
        since_days = compute_since_days(group_id, days)
        try:
            append_log(f"开始更新 {config['label']}，自上次更新起 {since_days} 天")
            old_rows = read_csv_rows(config["csv"])
            update_group(group_id, since_days)
            fresh = read_csv_rows(config["csv"])
            added = merge_csv_rows(fresh, config["csv"], existing=old_rows)
            total_new += added
            succeeded.append(group_id)
            set_last_update(group_id)
        except Exception as error:
            environment = os.environ.copy()
            environment["PYTHONUNBUFFERED"] = "1"
            ok, fallback_rows = rss_fallback_rows(config, since_days, environment)
            if ok:
                added = merge_csv_rows(fallback_rows, config["csv"])
                total_new += added
                succeeded.append(group_id)
                fallback_used.append(group_id)
                append_log(
                    f"{config['label']} 列表页故障，已用 RSS 兜底更新，新增 {added} 条"
                )
            else:
                failed.append(str(error))
                append_log(str(error))
    update_state(progress=95, message="正在整理数据…")
    refresh_recent_cache()
    if failed:
        detail = "；".join(failed)
        if succeeded:
            labels = [
                GROUPS[gid]["label"] + ("（RSS 兜底）" if gid in fallback_used else "")
                for gid in succeeded
            ]
            message = f"{'、'.join(labels)} 更新成功（新增 {total_new} 条）；" + detail
        else:
            message = f"更新失败；{detail}"
        update_state(
            status="error",
            message="部分更新失败" if succeeded else "更新失败",
            progress=0,
            error=message,
            finished_at=now_iso(),
        )
        return
    set_last_update()
    update_state(
        status="success",
        message=f"更新完成，新增 {total_new} 条",
        progress=100,
        results=[],
        workbook=str(latest_workbook()) if latest_workbook() else None,
        finished_at=now_iso(),
        error=None,
    )
    append_log(f"完成：新增 {total_new} 条")
    return


def start_update(days: int, group_ids: list[str]) -> bool:
    with STATE_LOCK:
        if STATE["status"] == "running":
            return False
        STATE.update(
            {
                "status": "running",
                "message": "正在准备更新…",
                "progress": 3,
                "days": days,
                "groups": group_ids,
                "results": [],
                "logs": [],
                "started_at": now_iso(),
                "finished_at": None,
                "error": None,
            }
        )
    threading.Thread(target=update_worker, args=(days, group_ids), daemon=True).start()
    return True


def initialize_state() -> None:
    load_versions()
    workbook = latest_workbook()
    update_state(
        workbook=str(workbook) if workbook is not None else None,
        results=[],
        message="正在读取数据…",
    )

    def preload() -> None:
        try:
            refresh_recent_cache()
            update_state(message="准备就绪")
        except Exception as error:
            append_log(str(error))
            update_state(message="准备就绪")

    threading.Thread(target=preload, daemon=True).start()


class AppHandler(BaseHTTPRequestHandler):
    server_version = "DMHYUpdater/1.0"

    def send_json(self, payload: dict, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def send_static(self, path: Path, content_type: str) -> None:
        if not path.exists():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        body = path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:
        route = unquote(urlparse(self.path).path)
        if route == "/":
            self.send_static(STATIC_DIR / "index.html", "text/html; charset=utf-8")
        elif route == "/styles.css":
            self.send_static(STATIC_DIR / "styles.css", "text/css; charset=utf-8")
        elif route == "/app.js":
            self.send_static(STATIC_DIR / "app.js", "text/javascript; charset=utf-8")
        elif route == "/api/status":
            self.send_json(state_snapshot())
        elif route == "/api/version":
            self.send_json(load_versions())
        elif route == "/api/recent":
            query = parse_qs(urlparse(self.path).query)
            try:
                days = min(max(int((query.get("days") or ["7"])[0]), 1), 365)
            except (ValueError, IndexError):
                days = 7
            raw_groups = (query.get("groups") or [""])[0].split(",")
            group_ids = [value for value in raw_groups if value in GROUPS]
            search_query = (query.get("q") or [""])[0].strip()
            fetch_all = (query.get("all") or ["0"])[0] in ("1", "true", "yes")
            cached = refresh_recent_cache()
            if cached.get("ready"):
                if search_query:
                    releases = search_releases(search_query, group_ids)
                elif fetch_all:
                    releases = cached_releases(group_ids)
                else:
                    releases = recent_releases(days, group_ids)
                releases = [
                    {key: value for key, value in item.items() if key != "published"}
                    for item in releases
                ]
            else:
                releases = []
            self.send_json(
                {
                    "days": days,
                    "groups": group_ids,
                    "q": search_query,
                    "all": fetch_all,
                    "total": len(releases),
                    "results": releases,
                    "workbook": str(cached["workbook"]) if cached["workbook"] is not None else None,
                }
            )
        elif route == "/api/search":
            query = parse_qs(urlparse(self.path).query)
            search_query = (query.get("q") or [""])[0].strip()
            group_value = (query.get("groups") or [""])[0]
            if group_value not in GROUPS:
                group_value = ""
            if not search_query:
                self.send_json({"q": "", "source": "site", "total": 0, "results": []})
                return
            try:
                self.send_json(site_search_cached(search_query, group_value))
            except Exception as error:
                self.send_json(
                    {"error": str(error)}, HTTPStatus.INTERNAL_SERVER_ERROR
                )
        elif route == "/api/workbook":
            today = datetime.now(SHANGHAI).strftime("%Y-%m-%d")
            target = DATA_DIR / f"DMHY更新表_{today}.xlsx"
            workbook = target if target.exists() else None
            if workbook is None:
                try:
                    workbook = generate_excel()
                    update_state(workbook=str(workbook))
                except Exception as error:
                    self.send_json(
                        {"error": f"生成表格失败：{error}"},
                        HTTPStatus.INTERNAL_SERVER_ERROR,
                    )
                    return
            if workbook is None or not workbook.exists():
                self.send_error(HTTPStatus.NOT_FOUND, "No workbook available")
                return
            body = workbook.read_bytes()
            encoded_name = workbook.name.encode("utf-8").hex()
            self.send_response(HTTPStatus.OK)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header("Content-Disposition", f"attachment; filename=DMHY-update.xlsx; x-filename-hex={encoded_name}")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        elif route == "/api/health":
            self.send_json({"ok": True, "demo": DEMO_MODE})
        else:
            self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        route = unquote(urlparse(self.path).path)
        if route != "/api/update":
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        try:
            length = min(int(self.headers.get("Content-Length", "0")), 16_384)
            payload = json.loads(self.rfile.read(length) or b"{}")
            days = int(payload.get("days", 7))
            group_ids = payload.get("groups", [])
            if not 1 <= days <= 365:
                raise ValueError("更新时间必须在 1 到 365 天之间")
            if not isinstance(group_ids, list) or not group_ids:
                raise ValueError("请至少选择一个字幕组")
            if any(group_id not in GROUPS for group_id in group_ids):
                raise ValueError("包含未知字幕组")
        except (ValueError, TypeError, json.JSONDecodeError) as error:
            self.send_json({"error": str(error)}, HTTPStatus.BAD_REQUEST)
            return
        if not start_update(days, group_ids):
            self.send_json({"error": "已有更新任务正在运行"}, HTTPStatus.CONFLICT)
            return
        self.send_json(state_snapshot(), HTTPStatus.ACCEPTED)

    def log_message(self, format_string: str, *args) -> None:
        return


def serve(port: int, open_browser: bool) -> None:
    server = None
    selected_port = port
    for candidate in range(port, port + 20):
        try:
            server = ThreadingHTTPServer(("127.0.0.1", candidate), AppHandler)
            selected_port = candidate
            break
        except OSError:
            continue
    if server is None:
        raise RuntimeError("无法找到可用的本地端口")
    url = f"http://127.0.0.1:{selected_port}"
    print(f"DMHY 追番助手已启动：{url}", flush=True)
    threading.Thread(target=initialize_state, daemon=True).start()
    if open_browser:
        threading.Timer(0.6, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever(poll_interval=0.25)
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


def main() -> int:
    global DEMO_MODE
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--port", type=int, default=8765)
    parser.add_argument("--open", action="store_true", help="open the browser after startup")
    parser.add_argument("--demo", action="store_true", help=argparse.SUPPRESS)
    args = parser.parse_args()
    DEMO_MODE = args.demo
    serve(args.port, args.open)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
